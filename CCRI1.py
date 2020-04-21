#!/usr/bin/env python3

"""
Open an existing workbook, copy one sheet
Copy data from one sheet to another
"""
from openpyxl import load_workbook

def main():
    #Create the Workbook object
    wprog = load_workbook("(U) ACC Progress Report v22-Tinker AFB (FOUO).xlsx")
    wb = load_workbook("(U) ACC estimated CCRI score (PhIV v1R4 31OCT17).xlsx")
    #Copy sheets
    sourceN = wb["EstimateNGradeACC"]
    sourceS = wb["EstimateSGradeACC"]

    new_sheet = wb.copy_worksheet(sourceN)
    new_sheet.title = "Copy of NGrade"
    
    new_sheet = wb.copy_worksheet(sourceS)
    new_sheet.title = "Copy of SGrade"

    # Check what sheets exists in the workbook
    print("Sheets in workbook:")
    for sheet in wb:
        print(sheet.title)

    wb.save("2.4_Hello_copies.xlsx")

if __name__ == "__main__":
    main()